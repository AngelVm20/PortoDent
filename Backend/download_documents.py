from sqlalchemy.orm import Session
import crud, schemas
from openpyxl import load_workbook
from datetime import date
from database import SessionLocal
from fastapi import Depends


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def consulta_to_xlsx(consulta: schemas.Consulta, paciente: schemas.Paciente, historia: schemas.HistoriaClinica):
    # Cargar la plantilla
    wb = load_workbook('template.xlsx')
    ws = wb.active

    # Suponemos que conoces las celdas exactas donde quieres insertar los datos
    #Esto es con respecto a los datos de cabecera de la historia clinica
    ws['B2'] = paciente.Nombre
    ws['F2'] = paciente.Apellido
    ws['J2'] = paciente.Sexo
    ws['K2'] = calcular_edad(paciente.FechaNacimiento)

    #ws['X1'] = consulta.ID_Consulta
    ws['L2'] = consulta.ID_HistoriaC
    
    #Esto es el campo ciclo de vida del paciente 
    ws['B4'] = consulta.RangoAños

    #Esto ya es con respecto al motico de la consulta
    ws['A7'] = consulta.MotivoC

    #Esto es con respecto al punto 2 enfermedad o problema actual
    ws['A10'] = consulta.EnfActual

    #Esto es con respecto al punto 3 antecedentes personales y familiares
    ws['B13'] = consulta.OpcionesAntecedentes
    ws['A14'] = consulta.Antecedentes

    #Esto es con respecto al punto 4 Signos vitales
    ws['C17'] = consulta.SignosVitales
    ws['F17'] = consulta.FrecuenciaCar
    ws['J17'] = consulta.Temperatura
    ws['L17'] = consulta.FrecuenciaRes

    #Esto es con respecto al punto 5 Examen del sistema estognatico
    ws['D20'] = consulta.OpcionesEstomatognatico
    ws['A21'] = consulta.ExamenEstomat

    #Esto es con respecto al punto 6 Odontograma
    ws['A24'] = consulta.Odontograma

    #Esto es con respecto al punto 7 indicadores de salud bucal

    #HIGIENE ORAL SIMPLIFICADA 
    #PIEZAS DENTALES
    ws['B35'] = consulta.PiezaDental16_17_55
    ws['B36'] = consulta.PiezaDental11_21_51
    ws['B37'] = consulta.PiezaDental26_27_65
    ws['B39'] = consulta.PiezaDental36_37_75
    ws['B40'] = consulta.PiezaDental31_41_71
    ws['B41'] = consulta.PiezaDental46_47_85


    #PLACA 
    ws['C35'] = consulta.Placa16_17_55
    ws['C36'] = consulta.Placa11_21_51
    ws['C37'] = consulta.Placa26_27_65
    ws['C39'] = consulta.Placa36_37_75
    ws['C40'] = consulta.Placa31_41_71
    ws['C41'] = consulta.Placa46_47_85
    ws['C42'] = consulta.TotalPlaca

    #CALCULO
    ws['D35'] = consulta.Calculo16_17_55
    ws['D36'] = consulta.Calculo11_21_51
    ws['D37'] = consulta.Calculo26_27_65
    ws['D39'] = consulta.Calculo36_37_75
    ws['D40'] = consulta.Calculo31_41_71
    ws['D41'] = consulta.Calculo46_47_85
    ws['D42'] = consulta.TotalCalculo


    #GINGIVITIS
    ws['E35'] = consulta.Gingivitis16_17_55
    ws['E36'] = consulta.Gingivitis11_21_51
    ws['E37'] = consulta.Gingivitis26_27_65
    ws['E39'] = consulta.Gingivitis36_37_75
    ws['E40'] = consulta.Gingivitis31_41_71
    ws['E41'] = consulta.Gingivitis46_47_85
    ws['E42'] = consulta.TotalGingivitis

    #ENF. PERIODONTAL
    ws['H34'] = consulta.EnfermedadPerio

    #MAL OCLUSION
    ws['J34'] = consulta.MalOclusion

    #FLUOROSIS
    ws['L34'] = consulta.Fluorosis



    #Esto es con respecto al punto 8 Indices CPO-ceo
    #Indice D
    ws['O34'] = consulta.IndiceC
    ws['P34'] = consulta.IndiceP
    ws['Q34'] = consulta.IndiceO
    ws['R34'] = consulta.TotalCPO

    #Indice d
    ws['O36'] = consulta.Indicedc
    ws['P36'] = consulta.Indicede
    ws['Q36'] = consulta.Indicedo
    ws['R36'] = consulta.Totalceo

    #Esto es con respecto al punto 10 planes de diagnostico, terapeutico y educacional
    ws['D45'] = consulta.OpcionPlan
    ws['A46'] = consulta.PlanDiagnostico

    #Esto es con respecto al punto 11 diagnostico
    ws['B49'] = consulta.Diagnostico
    ws['L49'] = consulta.Cie
    ws['Q49'] = consulta.PreoDef

    ws['B51'] = historia.FechaApertura
    ws['E51'] = consulta.FechaProximaConsulta

    #Esto es con respecto al punto 12 tratamiento
    ws['D55'] = consulta.Tratamientos
    ws['D57'] = consulta.Procedimientos
    ws['B56'] = consulta.FechaConsulta
    ws['D59'] = consulta.Prescripcion
    ws['B58'] = consulta.Codigo

    # Guardar el archivo temporalmente y retornar su nombre
    filename = f"{paciente.Nombre}{paciente.Apellido}_{paciente.Cedula}_{consulta.FechaConsulta}_consulta.xlsx"
    wb.save(filename)
    return filename


def calcular_edad(fecha_nacimiento):
    hoy = date.today()
    edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    return edad