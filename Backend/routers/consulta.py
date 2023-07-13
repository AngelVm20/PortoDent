from typing import List
from fastapi import Depends, APIRouter, HTTPException
from sqlalchemy.orm import Session
import crud, models, schemas
from database import SessionLocal, engine
from openpyxl import load_workbook
from fastapi.responses import FileResponse
from datetime import date


router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Consulta
@router.get("/consultas/{consulta_id}", response_model=schemas.Consulta)
def read_consulta(consulta_id: int, db: Session = Depends(get_db)):
    db_consulta = crud.get_consulta(db, consulta_id=consulta_id)
    if db_consulta is None:
        raise HTTPException(status_code=404, detail="Consulta no encontrada")
    return db_consulta

@router.get("/consultas/", response_model=List[schemas.Consulta])
def read_consultas(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    consultas = crud.get_consultas(db, skip=skip, limit=limit)
    return consultas



@router.post("/consultas/", response_model=schemas.Consulta)
def create_consulta(consulta: schemas.ConsultaCreate, db: Session = Depends(get_db)):
    db_consulta = crud.create_consulta(db=db, consulta=consulta, historia_id=consulta.ID_HistoriaC)
    return db_consulta


@router.put("/consultas/{consulta_id}", response_model=schemas.Consulta)
def update_consulta(consulta_id: int, consulta: schemas.Consulta, db: Session = Depends(get_db)):
    db_consulta = crud.update_consulta(db=db, consulta_id=consulta_id, consulta=consulta)
    if db_consulta is None:
        raise HTTPException(status_code=404, detail="Consulta no encontrada")
    return db_consulta

@router.get("/historias_clinicas/{historia_id}/consultas", response_model=List[schemas.Consulta])
def read_consultas_by_historia(historia_id: int, db: Session = Depends(get_db)):
    consultas = crud.get_consultas_by_historia(db, historia_id=historia_id)
    return consultas



@router.get("/consultas/{consulta_id}/download", response_model=schemas.Consulta)
def download_consulta(consulta_id: int, db: Session = Depends(get_db)):
    db_consulta = crud.get_consulta(db, consulta_id=consulta_id)
    if db_consulta is None:
        raise HTTPException(status_code=404, detail="Consulta no encontrada")
    
    db_paciente = db_consulta.historia_clinica.paciente
    
    filename = consulta_to_xlsx(db_consulta, db_paciente)

    return FileResponse(filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)


def calcular_edad(fecha_nacimiento):
    hoy = date.today()
    edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    return edad

def consulta_to_xlsx(consulta: schemas.Consulta, paciente: schemas.Paciente):
    # Cargar la plantilla
    wb = load_workbook('template.xlsx')
    ws = wb.active

    # Suponemos que conoces las celdas exactas donde quieres insertar los datos
    #Esto es con respecto a los datos de cabecera de la historia clinica
    ws['D2'] = paciente.Nombre
    ws['H2'] = paciente.Apellido
    ws['M2'] = paciente.Sexo
    ws['N2'] = calcular_edad(paciente.FechaNacimiento)

    ws['X1'] = consulta.ID_Consulta
    ws['O2'] = consulta.ID_HistoriaC
    
    #Esto es el campo ciclo de vida del paciente 
    ws['B4'] = consulta.RangoAños

    #Esto ya es con respecto al motico de la consulta
    ws['A7'] = consulta.MotivoC

    #Esto es con respecto al punto 2 enfermedad o problema actual
    ws['A10'] = consulta.EnfActual

    #Esto es con respecto al punto 3 antecedentes personales y familiares
    ws['B13'] = consulta.OpcionesAntecedentes
    ws['A14'] = consulta.Antecedentes

    #Esto es con respecto al punto 4 Signos vitales
    ws['B17'] = consulta.SignosVitales
    ws['F17'] = consulta.FrecuenciaCar
    ws['J17'] = consulta.Temperatura
    ws['O17'] = consulta.FrecuenciaRes

    #Esto es con respecto al punto 5 Examen del sistema estognatico
    ws['C20'] = consulta.OpcionesEstomatognatico
    ws['A21'] = consulta.ExamenEstomat

    #Esto es con respecto al punto 6 Odontograma
    ws['A24'] = consulta.Odontograma

    #Esto es con respecto al punto 7 indicadores de salud bucal
    ws['J48'] = consulta.IndicadoresSalud
    ws['J49'] = consulta.EnfermedadPerio
    ws['J50'] = consulta.MalOclusion
    ws['J51'] = consulta.Fluorosis

    #Esto es con respecto al punto 8 Indices CPO-ceo
    ws['S48'] = consulta.IndicesCPO

    #Esto es con respecto al punto 9 planes de diagnostico, terapeutico y educacional
    ws['D61'] = consulta.OpcionPlan
    ws['A62'] = consulta.PlanDiagnostico

    #Esto es con respecto al punto 11 diagnostico
    ws['A65'] = consulta.Diagnostico

    #Esto es con respecto al punto 12 tratamiento
    ws['D73'] = consulta.Tratamientos
    ws['I73'] = consulta.ProcedimientosE
    ws['B74'] = consulta.FechaConsulta
    ws['O73'] = consulta.Prescripcion
    ws['V73'] = consulta.Codigo


    # Guardar el archivo temporalmente y retornar su nombre
    filename = f"{paciente.Nombre}{paciente.Apellido}_{paciente.Cedula}_consulta.xlsx"
    wb.save(filename)
    return filename
